import csv
import os
from pathlib import Path
from typing import Any, Dict, List, Optional
import openpyxl

class SpreadsheetAnalyzer:
    """
    Structured perception analyzer for Excel (.xlsx) and CSV spreadsheets.
    Extracts sheets, audits formulas, computes summary statistics, and detects error cells.
    """

    @classmethod
    def analyze_file(
        cls,
        file_path: str,
        sheet_name: Optional[str] = None,
        max_preview_rows: int = 25
    ) -> Dict[str, Any]:
        """Performs structured analysis on a spreadsheet file."""
        resolved = Path(os.path.expanduser(file_path)).resolve()
        if not resolved.exists():
            return {"success": False, "error": f"File not found: {resolved}"}

        suffix = resolved.suffix.lower()
        if suffix in [".xlsx", ".xlsm", ".xltx"]:
            return cls._analyze_xlsx(resolved, sheet_name, max_preview_rows)
        elif suffix in [".csv", ".tsv"]:
            return cls._analyze_csv(resolved, max_preview_rows)
        else:
            return {"success": False, "error": f"Unsupported spreadsheet format: {suffix}"}

    @classmethod
    def _analyze_xlsx(
        cls,
        path: Path,
        target_sheet: Optional[str],
        max_preview_rows: int
    ) -> Dict[str, Any]:
        wb = None
        wb_data = None
        try:
            wb = openpyxl.load_workbook(str(path), data_only=False)
            sheet_names = wb.sheetnames

            active_sheet_name = target_sheet if target_sheet in sheet_names else sheet_names[0]
            sheet = wb[active_sheet_name]

            # Also load data-only copy for evaluated values
            wb_data = openpyxl.load_workbook(str(path), data_only=True)
            sheet_data = wb_data[active_sheet_name]

            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0

            formulas = []
            error_cells = []
            numeric_data: Dict[int, List[float]] = {}
            headers = []
            preview_rows = []

            for r_idx in range(1, max_row + 1):
                row_vals = []
                for c_idx in range(1, max_col + 1):
                    raw_val = sheet.cell(row=r_idx, column=c_idx).value
                    eval_val = sheet_data.cell(row=r_idx, column=c_idx).value
                    cell_coord = sheet.cell(row=r_idx, column=c_idx).coordinate

                    # Check for formulas
                    if isinstance(raw_val, str) and raw_val.startswith("="):
                        formulas.append({
                            "cell": cell_coord,
                            "formula": raw_val,
                            "evaluated_value": str(eval_val)
                        })

                    # Check for error values (#VALUE!, #REF!, #DIV/0!, etc.)
                    val_str = str(eval_val or "")
                    if any(err in val_str for err in ["#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"]):
                        error_cells.append({
                            "cell": cell_coord,
                            "error": val_str,
                            "formula": str(raw_val)
                        })

                    # Collect numeric columns
                    if isinstance(eval_val, (int, float)) and not isinstance(eval_val, bool):
                        if c_idx not in numeric_data:
                            numeric_data[c_idx] = []
                        numeric_data[c_idx].append(float(eval_val))

                    if r_idx <= max_preview_rows:
                        row_vals.append(str(eval_val) if eval_val is not None else "")

                if r_idx == 1:
                    headers = row_vals
                if r_idx <= max_preview_rows:
                    preview_rows.append(row_vals)

            # Compute column statistics
            col_stats = []
            for c_idx, nums in numeric_data.items():
                if nums:
                    col_name = headers[c_idx - 1] if c_idx - 1 < len(headers) and headers[c_idx - 1] else f"Column_{c_idx}"
                    col_stats.append({
                        "column_index": c_idx,
                        "column_name": col_name,
                        "count": len(nums),
                        "sum": round(sum(nums), 2),
                        "min": round(min(nums), 2),
                        "max": round(max(nums), 2),
                        "avg": round(sum(nums) / len(nums), 2)
                    })

            return {
                "success": True,
                "file": str(path),
                "total_sheets": len(sheet_names),
                "sheets": sheet_names,
                "active_sheet": active_sheet_name,
                "dimensions": {"rows": max_row, "columns": max_col},
                "headers": headers,
                "preview_rows": preview_rows,
                "formulas_detected": len(formulas),
                "formulas": formulas[:50],
                "errors_detected": len(error_cells),
                "error_cells": error_cells,
                "column_statistics": col_stats
            }
        except Exception as e:
            return {"success": False, "error": f"Failed to analyze Excel file: {str(e)}"}
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
            if wb_data:
                try:
                    wb_data.close()
                except Exception:
                    pass

    @classmethod
    def _analyze_csv(cls, path: Path, max_preview_rows: int) -> Dict[str, Any]:
        try:
            rows = []
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= max_preview_rows:
                        break
                    rows.append(row)

            headers = rows[0] if rows else []
            return {
                "success": True,
                "file": str(path),
                "format": "csv",
                "headers": headers,
                "preview_table": {
                    "total_preview_rows": len(rows),
                    "rows": rows
                }
            }
        except Exception as e:
            return {"success": False, "error": f"Failed to analyze CSV: {str(e)}"}
